# -*- coding: utf-8 -*-
# arabam_scraper_windows.py

import os, sys, time, random
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# --- Windows konsolda UTF-8 ---
try:
    if os.name == "nt":
        os.system("chcp 65001 >NUL")
        sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

BASE = "https://www.arabam.com"
BASE_LIST = f"{BASE}/ikinci-el/otomobil"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
)

# ── SVG parça id → Parça Adı ────────────────────────────────────────────────────
PART_NAME_MAP = {
    "B0701": "Sol Arka Kapı",
    "B0801": "Sol Ön Kapı",
    "B01201": "Ön Tampon",
    "B0301": "Sol Arka Çamurluk",
    "B01101": "Sol Ön Çamurluk",
    "B0901": "Sağ Ön Çamurluk",
    "B0401": "Sağ Arka Kapı",
    "B0501": "Sağ Ön Kapı",
    "B0101": "Sağ Arka Çamurluk",
    "B01001": "Motor Kaputu",
    "B0201": "Bagaj Kapağı",
    "B0601": "Tavan",
    "B01301": "Arka Tampon",
}

# ── Excel kolonları ─────────────────────────────────────────────────────────────
PROPERTY_COLUMNS = [
    "İlan No", "İlan Tarihi", "Marka", "Seri", "Model",
    "Yıl", "KM", "Vites", "Yakıt Tipi",
    "Kasa Tipi", "Renk", "Motor Hacmi", "Motor Gücü",
    "Çekiş", "Araç Durumu", "Boya-değişen", "Takas", "Kimden"
]
PART_COLUMNS = list(PART_NAME_MAP.values())
BASE_COLUMNS = ["Başlık", "Fiyat", "İl/İlçe", "Link"]
ALL_COLUMNS = BASE_COLUMNS + PROPERTY_COLUMNS + PART_COLUMNS

# ── Site etiketleri → Excel kolonu eşlemesi ─────────────────────────────────────
SITEKEY_TO_COL = {
    "İlan No": "İlan No",
    "İlan Tarihi": "İlan Tarihi",
    "Marka": "Marka",
    "Seri": "Seri",
    "Model": "Model",
    "Yıl": "Yıl",
    "Kilometre": "KM",
    "KM": "KM",
    "Vites Tipi": "Vites",
    "Vites": "Vites",
    "Yakıt Tipi": "Yakıt Tipi",
    "Kasa Tipi": "Kasa Tipi",
    "Renk": "Renk",
    "Motor Hacmi": "Motor Hacmi",
    "Motor Gücü": "Motor Gücü",
    "Çekiş": "Çekiş",
    "Araç Durumu": "Araç Durumu",
    "Boya-değişen": "Boya-değişen",
    "Takasa Uygun": "Takas",
    "Takas": "Takas",
    "Kimden": "Kimden",
}

def clean(s: str) -> str:
    return (s or "").replace("\xa0", " ").strip().strip('"').strip()

def map_props_to_columns(props: Dict[str, str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for k, v in props.items():
        key = clean(k).rstrip(":")
        col = SITEKEY_TO_COL.get(key)
        if col:
            out[col] = clean(v)
    return out

# ── Decoy arama listesi ─────────────────────────────────────────────────────────
DECOY_QUERIES = [
    "bmw 320i", "bmw 520i", "audi a4", "audi a3",
    "mercedes c200", "skoda superb", "volvo s60", "honda civic",
]

# ── Bekleme aralıkları ──────────────────────────────────────────────────────────
WAIT_ITEM_RANGE  = (3.0, 7.0)       # ilanlar arası (sn)
WAIT_PAGE_RANGE  = (60.0, 240.0)    # sayfa geçişi (sn): 1–4 dk
LONG_BREAK_RANGE = (180.0, 300.0)   # uzun mola (sn): 3–5 dk

# ── HTTP session (tek yerden) ───────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Connection": "keep-alive",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
})

def build_list_url(query: str) -> str:
    slug = "-".join(query.lower().split())
    return f"{BASE_LIST}/{slug}"

def turkish_day_month(now: datetime) -> str:
    ay_ad = {
        1:"ocak",2:"şubat",3:"mart",4:"nisan",5:"mayıs",6:"haziran",
        7:"temmuz",8:"ağustos",9:"eylül",10:"ekim",11:"kasım",12:"aralık"
    }
    return f"{now.day}{ay_ad[now.month]}"

def vehicle_token_from_query(q: str) -> str:
    return (q or "").strip().lower().split()[-1]

def fetch_html(url: str, retries: int = 2, timeout: int = 30) -> BeautifulSoup:
    last_err = None
    for attempt in range(retries + 1):
        try:
            r = SESSION.get(url, timeout=timeout)
            r.raise_for_status()
            return BeautifulSoup(r.text, "html.parser")
        except Exception as e:
            last_err = e
            # basit backoff
            time.sleep(1.5 * (attempt + 1))
    # son hata
    raise last_err

def append_row_to_xlsx(path: Path, row: Dict[str, str], columns: List[str]):
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        wb.save(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([row.get(c, "") for c in columns])
    wb.save(path)

# ================= Liste sayfası parse =================
def extract_listing_rows(list_soup: BeautifulSoup):
    table = list_soup.select_one("table#main-listing")
    if not table: return []
    tbody = table.find("tbody")
    if not tbody: return []
    return tbody.select("tr[id^='listing']")

def extract_listing_link_from_tr(tr) -> Optional[str]:
    a = tr.select_one("div.fade-out-content-wrapper a") or tr.find("a", href=True)
    if not a or not a.get("href"):
        return None
    href = a["href"]
    return href if href.startswith("http") else (BASE + href)

def find_next_page_url(list_soup: BeautifulSoup) -> Optional[str]:
    next_a = (list_soup.select_one("ul.pagination li a#pagingNext") or
              list_soup.select_one("ul.pagination li a#paging_next") or
              list_soup.select_one("a[rel='next']"))
    if next_a and next_a.get("href"):
        href = next_a["href"]
        return href if href.startswith("http") else (BASE + href)
    return None

# ================= Detay sayfası parse =================
def parse_title_and_price(detail_soup: BeautifulSoup) -> Tuple[str, str]:
    title = ""
    price = ""
    t_el = detail_soup.select_one("h1.product-name") or detail_soup.select_one("h1")
    p_el = (detail_soup.select_one(".product-price")
            or detail_soup.select_one("div.price")
            or detail_soup.select_one("span.price"))
    if t_el: title = clean(" ".join(t_el.stripped_strings))
    if p_el: price = clean(" ".join(p_el.stripped_strings))
    return title, price

def parse_location(detail_soup: BeautifulSoup) -> str:
    # Gerekirse geliştirilebilir (breadcrumbs veya ilan bilgilerinden çekilebilir)
    return ""

def parse_property_items(detail_soup: BeautifulSoup) -> Dict[str, str]:
    root = detail_soup.select_one("div.product-properties")
    if not root: return {}
    container = root.select_one("div.product-properties-details") or root
    items = container.select("div.property-item")
    out: Dict[str, str] = {}
    for it in items:
        k_el = it.select_one("div.property-key")
        v_el = it.select_one("div.property-value")
        key = clean(" ".join(k_el.stripped_strings)) if k_el else ""
        val = clean(" ".join(v_el.stripped_strings)) if v_el else ""
        if key:
            out[key] = val
    return out

def normalize_status(s: str) -> str:
    s = clean(s)
    mapping = {
        "Orijinal": "Orijinal",
        "Boyanmış": "Boyalı",
        "Boyalı": "Boyalı",
        "Lokal Boyalı": "Lokal Boyalı",
        "Değişmiş": "Değişmiş",
        "Belirtilmemiş": "Belirtilmemiş",
    }
    return mapping.get(s, s)

def parse_damage_map(detail_soup: BeautifulSoup) -> Dict[str, str]:
    root = detail_soup.select_one("#tab-damage-information .damage-information-container svg")
    if not root:
        root = (detail_soup.select_one("div.damage-information-container svg")
                or detail_soup.select_one("svg.db")
                or detail_soup.select_one("svg"))
    if not root:
        return {}
    # Tooltip attribute'ları farklı olabilir: uib-tooltip / data-original-title / title
    elements = root.select("[id]")
    parts: Dict[str, str] = {}
    for el in elements:
        pid = el.get("id") or ""
        status = (el.get("uib-tooltip")
                  or el.get("data-original-title")
                  or el.get("title")
                  or "")
        status = normalize_status(status)
        if not pid or not status:
            continue
        name = PART_NAME_MAP.get(pid, pid)
        parts[name] = status
    return parts

# ================= Decoy ziyaretleri ====================
def visit_decoy_pages(n: int = 1):
    k = max(1, min(2, n))
    chosen = random.sample(DECOY_QUERIES, k=k)
    print(f"🎭 Decoy ziyaretleri: {chosen}")
    for q in chosen:
        url = build_list_url(q)
        try:
            _ = fetch_html(url)
            wait = random.uniform(3.0, 8.0)
            print(f"   • {url} — bekleme {wait:.1f}s")
            time.sleep(wait)
        except Exception as e:
            print(f"   • Decoy hata ({type(e).__name__}): {e}")

# ================= Orkestra =============================
def scrape(query: str, max_pages: Optional[int] = None):
    list_url = build_list_url(query)

    # Dosya adı: {model}_{gün+ay}.xlsx  (örn. superb_11eylül.xlsx)
    now = datetime.now()
    veh = vehicle_token_from_query(query)
    dtr = turkish_day_month(now)
    out_path = Path(f"{veh}_{dtr}.xlsx")
    print(f"📄 Çıktı dosyası: {out_path}")

    current_url = list_url
    page_idx = 0

    pages_since_decoy = 0
    decoy_interval = random.randint(1, 2)       # 1–2 sayfada bir decoy
    next_long_break_at = random.randint(8, 12)  # 8–12 sayfada bir uzun mola

    while True:
        page_idx += 1
        print(f"\n📄 SAYFA {page_idx}: {current_url}")
        try:
            list_soup = fetch_html(current_url)
        except Exception as e:
            print(f"⚠ Liste sayfası hata ({type(e).__name__}): {e}")
            break

        trs = extract_listing_rows(list_soup)
        if not trs:
            print("⚠ İlan satırı bulunamadı; duruyorum.")
            break

        # Sayfadaki ilanları rastgele sırayla işle
        order = list(range(len(trs)))
        random.shuffle(order)

        for pos in order:
            tr = trs[pos]
            link = extract_listing_link_from_tr(tr)
            if not link:
                continue

            print(f"➡ İlan: {link}")
            try:
                detail_soup = fetch_html(link)
            except Exception as e:
                print(f"   • Detay hata ({type(e).__name__}): {e}")
                continue

            title, price = parse_title_and_price(detail_soup)
            props_raw = parse_property_items(detail_soup)
            props = map_props_to_columns(props_raw)
            parts = parse_damage_map(detail_soup)

            # Excel satırı
            row: Dict[str, str] = {
                "Başlık": title,
                "Fiyat": price,
                "İl/İlçe": parse_location(detail_soup),
                "Link": link,
            }
            for col in PROPERTY_COLUMNS:
                row[col] = props.get(col, "")

            for pname in PART_COLUMNS:
                row[pname] = parts.get(pname, "")

            try:
                append_row_to_xlsx(out_path, row, ALL_COLUMNS)
                print(f"💾 Kaydedildi: {row['Başlık'] or '(başlık yok)'}")
            except Exception as e:
                print(f"   • Excel yazma hatası ({type(e).__name__}): {e}")

            t = random.uniform(*WAIT_ITEM_RANGE)
            print(f"⏳ İlan arası bekleme: {t:.1f}s")
            time.sleep(t)

        # — sayfa sonu: decoy & uzun mola —
        pages_since_decoy += 1
        if pages_since_decoy >= decoy_interval:
            visit_decoy_pages(n=random.randint(1, 2))
            pages_since_decoy = 0
            decoy_interval = random.randint(1, 2)

        if page_idx >= next_long_break_at:
            long_wait = random.uniform(*LONG_BREAK_RANGE)
            print(f"🛋 Uzun mola: {long_wait/60:.1f} dk")
            time.sleep(long_wait)
            next_long_break_at = page_idx + random.randint(8, 12)

        # sonraki sayfa?
        next_url = find_next_page_url(list_soup)
        if not next_url:
            print("🏁 Sonraki sayfa yok, bitti.")
            break

        if max_pages and page_idx >= max_pages:
            print("🔚 max_pages sınırına ulaşıldı.")
            break

        page_wait = random.uniform(*WAIT_PAGE_RANGE)
        print(f"😴 Sayfa geçişi bekleme: {page_wait/60:.1f} dk")
        time.sleep(page_wait)
        current_url = next_url

    print(f"\n✅ Tamamlandı. Excel: {out_path}")

# ================= ÇALIŞTIR =============================
if __name__ == "__main__":
    if len(sys.argv) >= 2:
        q = sys.argv[1]
        try:
            max_pages = int(sys.argv[2]) if len(sys.argv) >= 3 else None
        except ValueError:
            max_pages = None
    else:
        q = input("Marka + model (örn: 'skoda superb'): ").strip()
        max_pages = None

    scrape(q, max_pages=max_pages)
