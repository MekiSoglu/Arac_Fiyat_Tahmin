# -*- coding: utf-8 -*-
# arabam_com_gundelik_playwright_hybrid.py
#
# - Playwright ile ilk sayfadan canonical + ?sort=startedAt.desc alınır
# - Requests ile tüm sayfalar çekilir, ilanlar SIRAYLA işlenir, 3–7 sn bekleme
# - Hedef gün mantığı (bugün/dün/ay-gün), eski ilk ilanda durur
# - Detay sayfasından property ve hasar haritası (uib-tooltip) okunur
# - Hasar haritasında "Belirtilmemiş"/boş -> "Orijinal"
# - ÇIKTI 1 (günlük): C:\Users\EXCALIBUR\Desktop\sahibinden\gundelik_{model}_ilanlari\{model}_{günay}.xlsx
# - ÇIKTI 2 (SABİT ŞABLON): C:\Users\EXCALIBUR\Desktop\sahibinden\gundelik_passat_ilanlari\full-passat.xlsx
#   * İki dosyaya da, KM ve Yıl filtreleri SADECE GEÇEN satırlar eklenir
#   * "Link"e göre tekilleştirme (aynı ilan ikinci kez yazılmaz)
#   * Şablon dosyada sadece var olan başlıklara göre yazılır

import re, time, argparse, random
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Set
from datetime import datetime, timedelta
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

# ========= Ayarlar =========
BASE      = "https://www.arabam.com"
BASE_LIST = f"{BASE}/ikinci-el/otomobil"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36")

HEADLESS = False
DEBUG    = True

FORCE_TEST_DATE = ""  # "2025-09-11" veya "11.09.2025" (boşsa sistem günü)

def log(*a):
    if DEBUG: print("[DBG]", *a, flush=True)

# ========= Sütunlar / MAP =========
PART_NAME_MAP = {
    "B0701": "Sol Arka Kapı","B0801": "Sol Ön Kapı","B01201": "Ön Tampon",
    "B0301": "Sol Arka Çamurluk","B01101": "Sol Ön Çamurluk","B0901": "Sağ Ön Çamurluk",
    "B0401": "Sağ Arka Kapı","B0501": "Sağ Ön Kapı","B0101": "Sağ Arka Çamurluk",
    "B01001": "Motor Kaputu","B0201": "Bagaj kapağı","B0601": "Tavan","B01301": "Arka Tampon",
}
PROPERTY_COLUMNS = [
    "İlan No","İlan Tarihi","Marka","Seri","Model","Yıl","KM","Vites","Yakıt Tipi",
    "Kasa Tipi","Renk","Motor Hacmi","Motor Gücü","Çekiş","Araç Durumu",
    "Boya-değişen","Takas","Kimden"
]
PART_COLUMNS = list(PART_NAME_MAP.values())
BASE_COLUMNS = ["Başlık","Fiyat","İl/İlçe","Link"]
ALL_COLUMNS  = BASE_COLUMNS + PROPERTY_COLUMNS + PART_COLUMNS

SITEKEY_TO_COL = {
    "İlan No":"İlan No","İlan Tarihi":"İlan Tarihi","Marka":"Marka",
    "Seri":"Seri","Model":"Model","Yıl":"Yıl",
    "Kilometre":"KM","KM":"KM",
    "Vites Tipi":"Vites","Vites":"Vites",
    "Yakıt Tipi":"Yakıt Tipi","Kasa Tipi":"Kasa Tipi","Renk":"Renk",
    "Motor Hacmi":"Motor Hacmi","Motor Gücü":"Motor Gücü","Çekiş":"Çekiş",
    "Araç Durumu":"Araç Durumu","Boya-değişen":"Boya-değişen",
    "Takasa Uygun":"Takas","Takas":"Takas","Kimden":"Kimden",
}

TR_MONTHS = "ocak şubat mart nisan mayıs haziran temmuz ağustos eylül ekim kasım aralık".split()
DATE_PAT  = re.compile(r"\b(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)(?:\s+\d{4})?\b", re.I)

# ========= Yardımcılar =========
def clean(s: Optional[str]) -> str:
    return (s or "").replace("\xa0"," ").strip().strip('"')

def parse_today_arg(s: Optional[str]) -> Optional[datetime]:
    if not s: return None
    s = s.strip()
    for fmt in ("%Y-%m-%d","%d.%m.%Y"):
        try: return datetime.strptime(s, fmt)
        except: pass
    raise ValueError("--today: YYYY-MM-DD ya da DD.MM.YYYY")

def resolved_today(today_arg: Optional[str]) -> datetime:
    if FORCE_TEST_DATE:
        return parse_today_arg(FORCE_TEST_DATE) or datetime.now()
    return parse_today_arg(today_arg) or datetime.now()

def build_list_url(query: str) -> str:
    slug = "-".join(query.lower().split())
    return f"{BASE_LIST}/{slug}"

def turkish_day_month(dt: datetime) -> str:
    ay = {1:"ocak",2:"şubat",3:"mart",4:"nisan",5:"mayıs",6:"haziran",7:"temmuz",8:"ağustos",9:"eylül",10:"ekim",11:"kasım",12:"aralık"}
    return f"{dt.day}{ay[dt.month]}"

def ensure_parent_dir(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)

def fetch_html(url: str) -> BeautifulSoup:
    r = requests.get(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "tr-TR,tr;q=0.9",
        "Cache-Control":"no-cache","Pragma":"no-cache",
    }, timeout=30)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")

def to_number(s: str) -> float:
    s = clean(s).replace(".","").replace(",",".")
    s = re.sub(r"[^0-9\.]","", s)
    try: return float(s) if s else float("nan")
    except: return float("nan")

def year_from_text(s: str) -> Optional[int]:
    m = re.search(r"\b(19|20)\d{2}\b", clean(s))
    return int(m.group(0)) if m else None

def should_keep_row(row: Dict[str,str], max_km: Optional[int], min_year: Optional[int]) -> bool:
    """Satır, KM ve Yıl filtrelerini geçiyorsa True."""
    if max_km is not None:
        km_val = to_number(row.get("KM",""))
        if km_val == km_val and km_val > max_km:  # NaN değil ve sınırdan büyük
            return False
    if min_year is not None:
        y = year_from_text(row.get("Yıl",""))
        if y is not None and y < min_year:
            return False
    return True

def read_headers(path: Path) -> List[str]:
    """Var olan dosyanın başlıklarını sırayla döndürür."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value if c.value is not None else "" for c in ws[1]]
    wb.close()
    return headers

def append_row_unique(path: Path, row: Dict[str,str], columns: List[str], unique_key: str = "Link"):
    """Dosyayı yoksa oluşturur. 'unique_key' (Link) zaten varsa eklemez."""
    ensure_parent_dir(path)
    if not path.exists():
        wb = Workbook(); ws = wb.active; ws.append(columns); wb.save(path)

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        key_idx = headers.index(unique_key) + 1
    except ValueError:
        key_idx = None

    existing: Set[str] = set()
    if key_idx:
        for r in range(2, ws.max_row + 1):
            existing.add(str(ws.cell(r, key_idx).value or ""))

    link_val = str(row.get(unique_key,""))
    if key_idx and link_val in existing:
        wb.close()
        return  # zaten var

    ws.append([row.get(c,"") for c in columns])
    wb.save(path); wb.close()

def append_row_to_existing_template(path: Path, row: Dict[str,str], unique_key: str = "Link"):
    """
    Var olan bir şablon dosyaya (oluşturmadan) yaz.
    Başlıkları şablondan okur; sadece o başlıklara karşılık gelen değerleri yazar.
    """
    if not path.exists():
        print(f"⚠ Şablon bulunamadı, atlanıyor: {path}")
        return

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Uniq kontrol
    try:
        key_idx = headers.index(unique_key) + 1
    except ValueError:
        key_idx = None

    existing: Set[str] = set()
    if key_idx:
        for r in range(2, ws.max_row + 1):
            existing.add(str(ws.cell(r, key_idx).value or ""))

    link_val = str(row.get(unique_key,""))
    if key_idx and link_val in existing:
        wb.close()
        return  # zaten var

    # Sadece şablondaki kolon sırasına göre yaz
    ws.append([row.get(h,"") for h in headers])
    wb.save(path); wb.close()

# ========= Tarih çözümleme =========
def date_from_text(text: str, reference_now: datetime) -> Optional[datetime.date]:
    t = (text or "").strip().lower()
    if not t: return None
    if "bugün" in t or "saat önce" in t or "dk önce" in t:
        return reference_now.date()
    if "dün" in t:
        return (reference_now - timedelta(days=1)).date()
    m = DATE_PAT.search(t)
    if m:
        gun = int(m.group(1)); ay_ad = m.group(2).lower()
        try: ay = TR_MONTHS.index(ay_ad)+1
        except ValueError: return None
        y_m = re.search(r"\b(\d{4})\b", t)
        yil = int(y_m.group(1)) if y_m else reference_now.year
        try: return datetime(yil, ay, gun).date()
        except ValueError: return None
    return None

def is_on_or_after_target(text: str, target_day: datetime, reference_now: datetime) -> Optional[bool]:
    d = date_from_text(text, reference_now)
    if d is None: return None
    return d >= target_day.date()

# ========= Liste / Detay parse =========
def parse_list_date_from_tr(tr) -> str:
    a = tr.select_one("td.listing-text.tac div.fade-out-content-wrapper a[href]")
    if a: return " ".join(a.stripped_strings).strip()
    t = tr.select_one("time")
    if t: return " ".join(t.stripped_strings).strip()
    for td in tr.select("td"):
        txt = " ".join(td.stripped_strings).strip().lower()
        if any(k in txt for k in ["bugün","dün","saat önce","dk önce"]): return txt
        if any(m in txt for m in TR_MONTHS): return txt
    return ""

def extract_listing_link_from_tr(tr) -> Optional[str]:
    a = tr.select_one("td.listing-text.tac div.fade-out-content-wrapper a[href]") \
        or tr.select_one("div.fade-out-content-wrapper a[href]") \
        or tr.find("a", href=True)
    if not a or not a.get("href"): return None
    href = a["href"]
    return href if href.startswith("http") else (BASE + href)

def extract_listing_rows_from_html(html: str):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.select_one("table#main-listing")
    tbody = table.find("tbody") if table else None
    return [] if not tbody else tbody.select("tr[id^='listing']")

def parse_property_items(detail_soup: BeautifulSoup) -> Dict[str,str]:
    root = detail_soup.select_one("div.product-properties")
    if not root: return {}
    cont = root.select_one("div.product-properties-details") or root
    items = cont.select("div.property-item"); out={}
    for it in items:
        k = it.select_one("div.property-key"); v = it.select_one("div.property-value")
        key = clean(" ".join(k.stripped_strings)) if k else ""
        val = clean(" ".join(v.stripped_strings)) if v else ""
        if key: out[key]=val
    return out

def parse_title_and_price(detail_soup: BeautifulSoup) -> Tuple[str,str]:
    t_el = detail_soup.select_one("h1.product-name") or detail_soup.select_one("h1")
    p_el = detail_soup.select_one(".product-price") or detail_soup.select_one("div.price") or detail_soup.select_one("span.price")
    return clean(" ".join(t_el.stripped_strings)) if t_el else "", clean(" ".join(p_el.stripped_strings)) if p_el else ""

def normalize_status(s: str) -> str:
    t = clean(s)
    if not t or t.lower().startswith("belirtilmemiş"):
        return "Orijinal"
    mapping = {
        "Orijinal": "Orijinal",
        "Boyanmış": "Boyalı",
        "Boyalı": "Boyalı",
        "Lokal Boyalı": "Lokal Boyalı",
        "Değişmiş": "Değişmiş",
    }
    return mapping.get(t, t)

def parse_damage_map(detail_soup: BeautifulSoup) -> Dict[str,str]:
    root = (detail_soup.select_one("#tab-damage-information .damage-information-container svg")
            or detail_soup.select_one("div.damage-information-container svg")
            or detail_soup.select_one("svg"))
    if not root: return {}
    parts={}
    for el in root.select("[id][uib-tooltip]"):
        pid = el.get("id") or ""; status = normalize_status(el.get("uib-tooltip") or "")
        if not pid or not status: continue
        name = PART_NAME_MAP.get(pid, pid); parts[name]=status
    return parts

def parse_location(detail_soup: BeautifulSoup) -> str:
    return ""

# ========= UI: ilk sayfa aç, canonical yakala =========
def try_close_cookie_banners(page):
    selectors = [
        "#onetrust-accept-btn-handler","button#onetrust-accept-btn-handler",
        "button:has-text('Kabul')","button:has-text('Tümünü Kabul Et')",
        "button:has-text('Kabul Et')","button:has-text('Accept')",
        "text=Tümünü kabul et","text=Kabul Et",
    ]
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=3000)
            page.click(sel)
            log("Cookie banner kapatıldı:", sel)
            return True
        except Exception:
            continue
    return False

def get_sorted_base_url_via_ui(query: str) -> Tuple[str, str]:
    list_url = build_list_url(query)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(user_agent=UA, java_script_enabled=True)
        page = context.new_page()
        log("Sayfa açılıyor:", list_url)
        page.goto(list_url, timeout=60000, wait_until="domcontentloaded")
        try_close_cookie_banners(page)
        try:
            canonical = page.eval_on_selector("link[rel='canonical']", "el => el.href")
        except Exception:
            canonical = None
        if not canonical:
            canonical = page.url
        purl = urlparse(canonical)
        qs = dict(parse_qsl(purl.query, keep_blank_values=True))
        qs["sort"] = "startedAt.desc"
        new_q = urlencode(qs, doseq=True)
        resolved = urlunparse((purl.scheme, purl.netloc, purl.path, "", new_q, ""))
        html = page.content()
        browser.close()
        return html, resolved

# ========= Orkestra =========
def scrape(query: str, max_km: Optional[int], min_year: Optional[int],
           today_arg: Optional[str], max_pages: Optional[int]=None):
    target_day = resolved_today(today_arg)

    # Model token ve günlük çıktı yolu
    model = query.split()[-1].lower()
    day_mon = turkish_day_month(target_day)

    # Günlük dosya: C:\Users\EXCALIBUR\Desktop\sahibinden\gundelik_{model}_ilanlari\{model}_{günay}.xlsx
    daily_dir  = Path(r"C:\Users\EXCALIBUR\Desktop\sahibinden\megane") / f"gundelik_{model}_ilanları"
    daily_path = daily_dir / f"{model}_{day_mon}.xlsx"

    # SABİT ŞABLON (mevcut olmalı): C:\Users\EXCALIBUR\Desktop\sahibinden\gundelik_passat_ilanlari\full-passat.xlsx
    fixed_full_path = Path(r"C:\Users\EXCALIBUR\Desktop\sahibinden\megane\gundelik_megane_ilanları\megane_12eylül.xlsx")

    print(f"📄 Günlük: {daily_path}")
    print(f"📦 Şablon (sabit): {fixed_full_path}")

    # İlk sayfa: UI → canonical → sorted base url
    _, base_sorted_url = get_sorted_base_url_via_ui(query)
    print(f"🔗 Taban URL (sorted): {base_sorted_url}")

    total_kept = 0
    page_idx = 0
    stop_all = False

    session = requests.Session()
    session.headers.update({
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "tr-TR,tr;q=0.9",
        "Cache-Control":"no-cache","Pragma":"no-cache",
    })

    def page_url(n: int) -> str:
        p = urlparse(base_sorted_url)
        qs = dict(parse_qsl(p.query, keep_blank_values=True))
        qs["sort"] = "startedAt.desc"
        if n > 1:
            qs["page"] = str(n)
        else:
            qs.pop("page", None)
        new_q = urlencode(qs, doseq=True)
        return urlunparse((p.scheme, p.netloc, p.path, "", new_q, ""))

    while True:
        page_idx += 1
        if max_pages and page_idx > max_pages:
            print("🔚 max_pages sınırı.")
            break

        url = page_url(page_idx)
        log("GET", url)
        r = session.get(url, timeout=30)
        if r.status_code != 200:
            print(f"⚠ Sayfa {page_idx} getirilemedi (HTTP {r.status_code}).")
            break
        html = r.text

        print(f"\n📄 SAYFA {page_idx}: {url}")
        trs = extract_listing_rows_from_html(html)
        log("Liste satır sayısı:", len(trs))
        if not trs:
            print("⚠ İlan satırı bulunamadı; duruyorum.")
            break

        for tr in trs:
            list_date = parse_list_date_from_tr(tr)
            cmp = is_on_or_after_target(list_date, target_day=target_day, reference_now=datetime.now())

            if cmp is False:
                print(f"⛔ Hedef günden eski ilan: '{list_date}'. Tarama bitiriliyor.")
                stop_all = True
                break

            link = extract_listing_link_from_tr(tr)
            if not link:
                continue

            # Gerekirse detayda tarih doğrulaması
            if cmp is None:
                try:
                    detail_check = fetch_html(link)
                except Exception as e:
                    print(f"⚠ Detay alınamadı ({e}). Atlanıyor.")
                    continue
                det_props_raw = parse_property_items(detail_check)
                det_props = {}
                for k, v in det_props_raw.items():
                    col = SITEKEY_TO_COL.get(clean(k).rstrip(":"))
                    if col: det_props[col] = clean(v)
                det_date_text = det_props.get("İlan Tarihi","")
                cmp_det = is_on_or_after_target(det_date_text, target_day=target_day, reference_now=datetime.now())
                if cmp_det is False:
                    print(f"⛔ Detayda hedef günden eski: '{det_date_text}'. Tarama bitiriliyor.")
                    stop_all = True
                    break
                elif cmp_det is None:
                    print(f"⚠ Detay tarih çözülemedi ('{det_date_text}'). Atlanıyor.")
                    continue

                title, price = parse_title_and_price(detail_check)
                parts = parse_damage_map(detail_check)
                row = {"Başlık": title, "Fiyat": price, "İl/İlçe": parse_location(detail_check), "Link": link}
                for c in PROPERTY_COLUMNS: row[c] = det_props.get(c, "")
                for pn in PART_COLUMNS:     row[pn] = parts.get(pn, "")

                # --- filtre + yaz ---
                if should_keep_row(row, max_km, min_year):
                    # Günlük dosyaya (gerekirse oluşturur)
                    append_row_unique(daily_path, row, ALL_COLUMNS, unique_key="Link")
                    # SABİT ŞABLONA (oluşturmadan, başlığa göre eşleştirerek)
                    append_row_to_existing_template(fixed_full_path, row, unique_key="Link")
                    total_kept += 1
                    print(f"💾 Kaydedildi (detay-doğrulandı): {row['Başlık'] or '(başlık yok)'}")
                else:
                    print("🚫 Filtre nedeniyle eklenmedi.")
                time.sleep(random.uniform(3.0, 7.0))
                continue

            # cmp True: normal akış
            try:
                detail_soup = fetch_html(link)
            except Exception as e:
                print(f"   • Detay hata: {e}")
                continue

            title, price = parse_title_and_price(detail_soup)
            props_raw = parse_property_items(detail_soup)
            props = {}
            for k, v in props_raw.items():
                col = SITEKEY_TO_COL.get(clean(k).rstrip(":"))
                if col: props[col] = clean(v)
            parts = parse_damage_map(detail_soup)

            row = {"Başlık": title, "Fiyat": price, "İl/İlçe": parse_location(detail_soup), "Link": link}
            for c in PROPERTY_COLUMNS: row[c] = props.get(c, "")
            for pn in PART_COLUMNS:     row[pn] = parts.get(pn, "")

            if should_keep_row(row, max_km, min_year):
                append_row_unique(daily_path, row, ALL_COLUMNS, unique_key="Link")
                append_row_to_existing_template(fixed_full_path, row, unique_key="Link")
                total_kept += 1
                print(f"💾 Kaydedildi: {row['Başlık'] or '(başlık yok)'}")
            else:
                print("🚫 Filtre nedeniyle eklenmedi.")

            time.sleep(random.uniform(3.0, 7.0))

        if stop_all:
            break

    print(f"\n✅ Alınan ve filtreyi geçen ilan: {total_kept}")
    print(f"📄 Günlük dosya: {daily_path}")
    print(f"📦 Şablon dosya: {fixed_full_path}")

# ========= CLI =========
def ask_int(prompt: str) -> Optional[int]:
    s = input(prompt).strip()
    if not s: return None
    try: return int(s)
    except: return None

def main():
    ap = argparse.ArgumentParser(description="arabam.com — canonical + sort=startedAt.desc ile çek, filtrele, iki dosyaya yaz (günlük + sabit şablon).")
    ap.add_argument("--query", type=str, default=None, help="Arama (örn: 'volkswagen passat' ya da 'passat')")
    ap.add_argument("--min-year", type=int, default=None, help="En düşük model yılı")
    ap.add_argument("--max-km", type=int, default=None, help="En yüksek KM")
    ap.add_argument("--today", type=str, default=None, help="Hedef gün (YYYY-MM-DD veya DD.MM.YYYY)")
    ap.add_argument("--max-pages", type=int, default=None, help="Maks. sayfa (opsiyonel)")
    args = ap.parse_args()

    q = args.query or input("Marka + model: ").strip()
    min_year = args.min_year if args.min_year is not None else ask_int("En düşük model yılı (boş): ")
    max_km   = args.max_km   if args.max_km   is not None else ask_int("En yüksek KM (boş): ")

    scrape(q, max_km=max_km, min_year=min_year, today_arg=args.today, max_pages=args.max_pages)

if __name__ == "__main__":
    main()
